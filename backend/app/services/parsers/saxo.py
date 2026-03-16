"""
Saxo Bank XLSX parser.

Parses the Saxo Bank "Transacties" Excel export which contains 3 sheets:
- Transacties: Main transaction list (trades, dividends, cash, corporate actions)
- _Transacties: Detailed trade info (exact price, quantity, open/close)
- Bookings: Detailed cost breakdown per transaction (commission, tax, etc.)
"""
import re
from datetime import datetime
from io import BytesIO
from typing import Optional

import openpyxl

from .base import (
    BaseParser,
    ParseResult,
    ParsedTransaction,
    ParsedDividend,
    ParsedCashTransaction,
    ParsedStock,
)

# Saxo symbol -> Yahoo Finance ticker mapping
# Saxo uses exchange-specific symbols (e.g., BRYN:xetr)
SAXO_TO_YAHOO_TICKER = {
    "ABI:xbru": "ABI.BR",
    "AD:xams": "AD.AS",
    "BRYN:xetr": "BRK-B",
    "CPINV:xbru": "CPINV.BR",
    "ECL:xnys": "ECL",
    "ENGI:xpar": "ENGI.PA",
    "EQQQ:xpar": "EQQQ.DE",
    "IH2O:xmil": "IH2O.MI",
    "IBGX:xams": "IBGX.AS",
    "IUSA:xams": "IUSA.AS",
    "IWDA:xams": "IWDA.AS",
    "INRG:xmil": "INRG.MI",
    "LOCK-2673:PAR": "ENGI.PA",
    "LSPX:xlon": "SPXD.L",
    "MSFT:xnas": "MSFT",
    "RDSa:xams": "SHEL.AS",
    "TSLA:xnas": "TSLA",
    "ISPA:xams": "ISPA.AS",
    "XYL:xnys": "XYL",
}

# Country detection from ISIN prefix
ISIN_COUNTRY_MAP = {
    "US": "Verenigde Staten",
    "BE": "België",
    "NL": "Nederland",
    "FR": "Frankrijk",
    "DE": "Duitsland",
    "IE": "Ierland",
    "GB": "Verenigd Koninkrijk",
    "LU": "Luxemburg",
}


def _parse_action_qty_price(action_text: str) -> tuple[str, Optional[float], Optional[float]]:
    """
    Parse Saxo 'Acties' field like 'Koop 1 @ 405.85 USD' or 'Verkoop -3 @ 305.00 USD'.
    Returns (action_type, quantity, price).
    """
    if not action_text:
        return ("UNKNOWN", None, None)

    # Match patterns like "Koop 1 @ 114.09 EUR" or "Verkoop -16 @ 4,902.80 GBp"
    match = re.match(
        r'(Koop|Verkoop)\s+(-?\d[\d.,]*)\s+@\s+([\d.,]+)\s+(\w+)',
        action_text
    )
    if match:
        action = "BUY" if match.group(1) == "Koop" else "SELL"
        qty_str = match.group(2).replace(",", "")
        price_str = match.group(3).replace(",", "")
        return (action, abs(float(qty_str)), float(price_str))

    return ("UNKNOWN", None, None)


def _country_from_isin(isin: str) -> str:
    """Determine country from ISIN prefix."""
    if isin and len(isin) >= 2:
        return ISIN_COUNTRY_MAP.get(isin[:2], "Onbekend")
    return "Onbekend"


def _saxo_to_yahoo(saxo_symbol: str) -> Optional[str]:
    """Convert Saxo symbol to Yahoo Finance ticker."""
    return SAXO_TO_YAHOO_TICKER.get(saxo_symbol)


def _detect_asset_type(instrument_type: str) -> str:
    """Map Saxo instrument type to app asset type."""
    if instrument_type and instrument_type.lower() in ("etf", "etc", "etn"):
        return "STOCK"  # ETFs are treated as STOCK in the app
    return "STOCK"


class SaxoParser(BaseParser):
    """Parser for Saxo Bank XLSX transaction exports."""

    def supported_extensions(self) -> list[str]:
        return [".xlsx"]

    def parse(self, file_content: BytesIO, filename: str) -> ParseResult:
        result = ParseResult(broker="Saxo")

        wb = openpyxl.load_workbook(file_content, read_only=True, data_only=True)

        # Parse main Transacties sheet
        if "Transacties" not in wb.sheetnames:
            result.warnings.append("Sheet 'Transacties' niet gevonden in het bestand.")
            return result

        # Build trade details lookup from _Transacties sheet
        trade_details = {}
        if "_Transacties" in wb.sheetnames:
            trade_details = self._parse_trade_details(wb["_Transacties"])

        # Build bookings lookup from Bookings sheet
        bookings = {}
        if "Bookings" in wb.sheetnames:
            bookings = self._parse_bookings(wb["Bookings"])

        # Parse main transactions
        self._parse_transactions(
            wb["Transacties"], trade_details, bookings, result
        )

        wb.close()
        return result

    def _parse_trade_details(self, ws) -> dict:
        """
        Parse _Transacties sheet for exact trade details.
        Returns dict keyed by Transactie-ID.
        """
        details = {}
        headers = None

        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = [str(h).strip() if h else "" for h in row]
                continue

            row_dict = dict(zip(headers, row))
            tx_id = row_dict.get("Transactie-ID")
            if tx_id:
                details[str(tx_id)] = {
                    "quantity": row_dict.get("Traded\xa0Quantity") or row_dict.get("Traded Quantity"),
                    "price": row_dict.get("Prijs"),
                    "open_close": row_dict.get("Openen/sluiten"),
                    "trade_type": row_dict.get("Trade\xa0Event\xa0Type") or row_dict.get("Trade Event Type"),
                }
        return details

    def _parse_bookings(self, ws) -> dict:
        """
        Parse Bookings sheet for cost breakdowns.
        Returns dict keyed by Bk Record Id -> list of booking entries.
        """
        bookings = {}
        headers = None

        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = [str(h).strip() if h else "" for h in row]
                continue

            row_dict = dict(zip(headers, row))
            bk_id = row_dict.get("Bk\xa0Record\xa0Id") or row_dict.get("Bk Record Id")
            if bk_id:
                bk_key = str(bk_id)
                if bk_key not in bookings:
                    bookings[bk_key] = []
                bookings[bk_key].append({
                    "amount_type": row_dict.get("Amount Type", ""),
                    "amount": row_dict.get("Boekingsbedrag", 0) or 0,
                    "conversion_cost": row_dict.get("Conversion cost", 0) or 0,
                    "ex_date": row_dict.get("Ex-datum"),
                    "eligible_quantity": row_dict.get("Eligible quantity"),
                    "dividend_per_share": row_dict.get("Dividend per share"),
                    "tax_percentage": row_dict.get("Tax\xa0Percentage") or row_dict.get("Tax Percentage"),
                })
        return bookings

    def _parse_transactions(self, ws, trade_details: dict, bookings: dict,
                            result: ParseResult):
        """Parse the main Transacties sheet."""
        headers = None
        seen_stocks = {}  # ticker -> ParsedStock

        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = [str(h).strip() if h else "" for h in row]
                continue

            row_dict = dict(zip(headers, row))

            tx_type = row_dict.get("Transactietype", "")
            action = row_dict.get("Acties", "")
            instrument = row_dict.get("Instrument")
            saxo_symbol = row_dict.get("Instrumentsymbool")
            isin = row_dict.get("Instrument ISIN") or ""
            instrument_currency = row_dict.get("Instrumentvaluta", "EUR")
            account_id = row_dict.get("Rekening-ID", "")
            tx_id = row_dict.get("Transactie-ID")
            bk_record_id = row_dict.get("Bk\xa0Record\xa0Id") or row_dict.get("Bk Record Id")
            tx_date = row_dict.get("Transactiedatum")
            booking_amount = row_dict.get("Boekingsbedrag", 0) or 0
            fx_rate = row_dict.get("Omrekeningskoers", 1) or 1
            total_costs = row_dict.get("Totale kosten", 0) or 0

            if not tx_date:
                result.skipped_rows += 1
                continue

            # Convert date
            if isinstance(tx_date, datetime):
                date_str = tx_date.strftime("%Y-%m-%d")
            else:
                result.skipped_rows += 1
                continue

            # Determine account currency from account ID
            account_currency = "EUR"
            if account_id and "USD" in account_id:
                account_currency = "USD"

            # ============================================================
            # TRADES: Koop / Verkoop
            # ============================================================
            if tx_type == "Transactie" and action and ("Koop" in action or "Verkoop" in action):
                self._handle_trade(
                    row_dict, date_str, tx_id, bk_record_id,
                    action, instrument, saxo_symbol, isin,
                    instrument_currency, account_currency,
                    fx_rate, total_costs,
                    trade_details, bookings, seen_stocks, result
                )

            # ============================================================
            # CORPORATE ACTIONS: Dividends, withholding tax
            # ============================================================
            elif tx_type == "Corporate action":
                self._handle_corporate_action(
                    row_dict, date_str, action, instrument, saxo_symbol, isin,
                    instrument_currency, booking_amount, fx_rate,
                    bk_record_id, bookings, result
                )

            # ============================================================
            # CASH: Deposits, withdrawals
            # ============================================================
            elif tx_type in ("Storting/opname", "Cashbedrag"):
                self._handle_cash(
                    row_dict, date_str, action, account_currency,
                    booking_amount, fx_rate, total_costs, result
                )

            else:
                result.skipped_rows += 1

        # Add unique stocks to result
        result.stocks = list(seen_stocks.values())

    def _handle_trade(self, row_dict, date_str, tx_id, bk_record_id,
                      action, instrument, saxo_symbol, isin,
                      instrument_currency, account_currency,
                      fx_rate, total_costs,
                      trade_details, bookings, seen_stocks, result):
        """Process a buy/sell trade row."""
        parsed_action, qty, price = _parse_action_qty_price(action)
        if not qty or not price:
            result.warnings.append(f"Kon transactie niet parsen: {action}")
            result.skipped_rows += 1
            return

        # Enrich from _Transacties if available
        if tx_id and str(tx_id) in trade_details:
            detail = trade_details[str(tx_id)]
            if detail.get("quantity"):
                qty = abs(float(detail["quantity"]))
            if detail.get("price"):
                price = float(detail["price"])

        # Determine fees and taxes from Bookings
        fees = 0.0
        taxes = 0.0
        if bk_record_id and str(bk_record_id) in bookings:
            for bk in bookings[str(bk_record_id)]:
                amt_type = bk.get("amount_type", "").lower()
                amt = abs(bk.get("amount", 0) or 0)
                if "commissie" in amt_type:
                    fees += amt
                elif "beurstaks" in amt_type:
                    taxes += amt
                elif "financiële transactiebelasting" in amt_type:
                    taxes += amt
                elif "beurskosten" in amt_type:
                    fees += amt
                elif "conversion cost" in amt_type:
                    fees += abs(bk.get("conversion_cost", 0) or 0)
        elif total_costs:
            # Fallback: use total_costs as fees
            fees = abs(total_costs)

        # Map ticker
        yahoo_ticker = _saxo_to_yahoo(saxo_symbol) if saxo_symbol else None
        app_ticker = yahoo_ticker or (saxo_symbol.split(":")[0] if saxo_symbol else isin)

        # Handle GBP pence (GBp) - convert to GBP
        if instrument_currency == "GBP" and price > 100:
            # Saxo reports London stocks in pence, convert to pounds
            price = price / 100.0

        # Determine the exchange rate for the transaction
        # For EUR account trades in EUR instruments, rate is 1.0
        # For USD account trades, fx_rate is the EUR/USD rate at time of trade
        exchange_rate = float(fx_rate) if fx_rate else 1.0

        tx = ParsedTransaction(
            date=date_str,
            broker="Saxo",
            transaction_type=parsed_action,
            name=instrument or "Onbekend",
            ticker=app_ticker,
            isin=isin,
            quantity=qty,
            price_per_share=price,
            currency=instrument_currency or "EUR",
            fees=round(fees, 2),
            taxes=round(taxes, 2),
            exchange_rate=exchange_rate,
            fees_currency=account_currency,
            notes=f"Saxo {account_currency} rekening" if account_currency != "EUR" else None,
            source_id=str(tx_id) if tx_id else None,
        )
        result.transactions.append(tx)

        # Track stock info
        if app_ticker and app_ticker not in seen_stocks:
            seen_stocks[app_ticker] = ParsedStock(
                ticker=app_ticker,
                isin=isin,
                name=instrument or "Onbekend",
                asset_type=_detect_asset_type(row_dict.get("Type", "")),
                currency=instrument_currency or "EUR",
                yahoo_ticker=yahoo_ticker,
                country=_country_from_isin(isin),
            )

    def _handle_corporate_action(self, row_dict, date_str, action, instrument,
                                  saxo_symbol, isin, instrument_currency,
                                  booking_amount, fx_rate, bk_record_id, bookings,
                                  result):
        """Process corporate actions (dividends, withholding tax, etc.)."""
        action_lower = action.lower() if action else ""

        # Map ticker
        yahoo_ticker = _saxo_to_yahoo(saxo_symbol) if saxo_symbol else None
        app_ticker = yahoo_ticker or (saxo_symbol.split(":")[0] if saxo_symbol else isin)

        # ---- Cash Dividend ----
        if "cashdividend" in action_lower or "herbeleggingsdividend" in action_lower:
            bruto = abs(float(booking_amount)) if booking_amount else 0
            if bruto == 0:
                return

            # Look for associated withholding tax in bookings
            withheld_tax = 0.0
            if bk_record_id and str(bk_record_id) in bookings:
                for bk in bookings[str(bk_record_id)]:
                    amt_type = bk.get("amount_type", "").lower()
                    if "bronbelasting" in amt_type or "roerende voorheffing" in amt_type:
                        withheld_tax += abs(bk.get("amount", 0) or 0)

            net = round(bruto - withheld_tax, 2) if withheld_tax else bruto

            div = ParsedDividend(
                ticker=app_ticker or "UNKNOWN",
                isin=isin,
                ex_date=date_str,
                bruto_amount=round(bruto, 2),
                currency=instrument_currency or "EUR",
                withheld_tax=round(withheld_tax, 2),
                net_amount=net,
                received=True,
                notes=f"{'Herbeleggingsdividend' if 'herbelegging' in action_lower else 'Cashdividend'} via Saxo",
            )
            result.dividends.append(div)

        # ---- Roerende voorheffing (Belgian withholding tax - standalone) ----
        elif "roerende voorheffing" in action_lower:
            # This is a standalone RV entry - usually a refund or separate tax event
            # These are already included in the booking breakdown of dividends
            # but some appear as standalone corporate actions
            amount = float(booking_amount) if booking_amount else 0
            if amount > 0:
                # Positive = tax refund, treat as dividend income
                div = ParsedDividend(
                    ticker=app_ticker or "UNKNOWN",
                    isin=isin,
                    ex_date=date_str,
                    bruto_amount=round(amount, 2),
                    currency=instrument_currency or "EUR",
                    withheld_tax=0,
                    net_amount=round(amount, 2),
                    received=True,
                    notes="Roerende voorheffing (terugbetaling) via Saxo",
                )
                result.dividends.append(div)
            # Negative amounts are tax payments already captured in dividend entries

        # ---- Kapitaaluitkering (capital return) ----
        elif "kapitaaluitkering" in action_lower:
            amount = abs(float(booking_amount)) if booking_amount else 0
            if amount > 0:
                div = ParsedDividend(
                    ticker=app_ticker or "UNKNOWN",
                    isin=isin,
                    ex_date=date_str,
                    bruto_amount=round(amount, 2),
                    currency=instrument_currency or "EUR",
                    withheld_tax=0,
                    net_amount=round(amount, 2),
                    received=True,
                    notes="Kapitaaluitkering via Saxo",
                )
                result.dividends.append(div)

        # Skip: Pari-Passu, Aandelensplitsing, Waardeloos, Notering ex-dividend
        else:
            result.skipped_rows += 1

    def _handle_cash(self, row_dict, date_str, action, account_currency,
                     booking_amount, fx_rate, total_costs, result):
        """Process cash transactions (deposits, withdrawals, interest, fees)."""
        action_lower = action.lower() if action else ""
        amount = float(booking_amount) if booking_amount else 0

        if "storting" in action_lower:
            tx_type = "DEPOSIT"
        elif "opname" in action_lower:
            tx_type = "WITHDRAWAL"
        elif "debetrente" in action_lower or "creditrente" in action_lower:
            tx_type = "INTEREST"
        elif "commissie" in action_lower:
            tx_type = "FEE"
        else:
            result.skipped_rows += 1
            return

        # Determine source currency for FX conversions
        source_amount = None
        source_currency = None
        exchange_rate = None

        valuta = row_dict.get("Valuta", "")
        boek_valuta = row_dict.get("_Valuta", "")
        if valuta and boek_valuta and valuta != boek_valuta:
            source_amount = row_dict.get("Boekingsbedrag")
            source_currency = boek_valuta
            exchange_rate = float(fx_rate) if fx_rate else None

        cash = ParsedCashTransaction(
            date=date_str,
            broker="Saxo",
            transaction_type=tx_type,
            amount=round(amount, 2),
            currency=account_currency,
            source_amount=round(source_amount, 2) if source_amount else None,
            source_currency=source_currency if source_currency else None,
            exchange_rate=exchange_rate,
            notes=action if action else None,
        )
        result.cash_transactions.append(cash)
